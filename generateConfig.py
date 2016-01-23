#!/usr/bin/env python
# -*- coding: utf8 -*-
import openpyxl
import json
import re

XLSfile = 'Seredniomisyachna_temper_vodi_нов.xlsx'
configFile = "configure.json"

def generate_file(confFile=configFile,XLS=XLSfile):
    listSHeetes = input("Please,choose massive of sheetes you want to work with.\nFor example,type: [5:11] or [5:] or [:11]  or just press ENTER to work with all sheetes.")
    wb = openpyxl.load_workbook(XLSfile)
    allSheetes = eval("wb.get_sheet_names()" + listSHeetes)
    print(allSheetes)
    print(len(allSheetes))
    with open(configFile,"w") as f:
        dataDict = {}
        for sheet in allSheetes:
            sht = wb.get_sheet_by_name(sheet)
            try:
                index_obj = re.search(r'[0-9]{5}',str(sht.cell(row=2,column=1).value))
                index = int(index_obj.group(0))
                chapterOfSheet = str(sht.cell(row=2,column=1).value)[index_obj.end():].rstrip().replace(' ','')
            except AttributeError:
                index = None
                chapterOfSheet = str(sht.cell(row=2,column=1).value).strip().replace(' ','')
            print(chapterOfSheet)
            oneSheet = {sheet:[index, chapterOfSheet]}
            dataDict.update(oneSheet)
        json.dump(dataDict, f, ensure_ascii=False, indent=4)

generate_file()