#!/usr/bin/env python
# -*- coding: utf8 -*-
import json
from difflib import SequenceMatcher
import csv
import re
from tkinter import filedialog
import tkinter as tk
import os
from .parsedoc import ParseOneDoc as pod
import openpyxl

opts = {'filetypes' : [('XLSX files','.xlsx'), ('All files', '.*')], 'title' : 'Select XLSX file with your sheetes, where you should insert data'}

def load_json_conf(confFile):
	with open(confFile) as f:
		jsonData = json.load(f)
		return jsonData

def seeDiff(jsonData,listOfPosts,ratioOfDiff=2.55):
    for postConvertFrom in listOfPosts:
        Flag = True
        for sheet in jsonData:
            postNameInSheetOrigin = jsonData[sheet][1]
            sheetTuple = tuple(re.sub(r'[-_]',' ',sheet).split())
            try:
                postNameInSheet = re.sub(r'[-_]',' ',postNameInSheetOrigin[postNameInSheetOrigin.index("р.") + 2:]).split()
            except ValueError:
                postNameInSheet = re.sub(r'[-_]',' ',postNameInSheetOrigin).split()
            rmList = []
            rnList = []
            for i,chunk in enumerate(postConvertFrom[0]):
                try:
                    m0 = SequenceMatcher(None,chunk,postNameInSheet[i])
                    rmList.append(m0.ratio())
                except IndexError:
                    pass
                try:
                    n0 = SequenceMatcher(None,chunk,sheetTuple[i])
                    rnList.append(n0.ratio())
                except IndexError:
                    pass
            rn = sum(rnList)
            rm = sum(rmList)
            mainRatio = rm + rn
            # if mainRatio > 2:
            #     print(postConvertFrom,"------","===>",mainRatio, postNameInSheetOrigin)    #('Деркул', 'смтБiловодськ') ------ ===> 0.0
            if mainRatio > ratioOfDiff:
                # print(("RATIO IS ", mainRatio, sheet, postConvertFrom[1]))
                Flag = False
                yield (sheet, postNameInSheetOrigin ,postConvertFrom[1], postConvertFrom[2], postConvertFrom[3])
        if Flag:
            print ("-----------------------------------------------")
            print ("Not incuded post names:    {0}\n<=======>{1}".format(postConvertFrom[1],postConvertFrom[2]))
            print ("-----------------------------------------------")

def write_posts_names_to_config(sheetesAndPosts, jsonData, jsonFile):              # sheetesAndPosts varieble must be a generator
    for pair in sheetesAndPosts:
    	if pair[2] not in jsonData[pair[0]]:
    		jsonData[pair[0]].append(pair[2])
    		#print(pair[2],pair[3])
    with open(jsonFile,"w") as f:
    	#print(jsonData)
    	json.dump(jsonData, f, ensure_ascii=False, indent=4)
    	# for data in jsonData:
    	# 	print(data)
    	# print(len(jsonData))


def walking(root):
    for folder in os.listdir(root):
        try:
            folderInteger = int(folder)
        except ValueError:
        	continue
        for dirPath, dirs, files in os.walk(os.path.join(root,folder)):
            files.sort()
            for fileSingle in files:
                if fileSingle[-3:] == 'xls' or fileSingle[-4:] == 'xlsx':
                    #print ('---PASS',fileSingle)
                    pass
                else:
                    yield (os.path.join(dirPath,fileSingle),folderInteger)

class createConf(pod):

    def __init__(self):
        pod.__init__(self)
        #self.filenameGenerator = filenameGenerator

    def parseDoc(self, filenameGenerator, indexList):
        self.filenameGenerator = filenameGenerator
        self.indexList = indexList
        for filename,year in self.filenameGenerator:
            with open(filename, 'rb') as f:
                counter = 0
                mark = False
                for r in f:
                    if not mark:
                        if not re.search(self.table_number_pattern, r.decode('ibm866').lower()) and not re.search(self.table_name_pattern,r.decode('ibm866').lower()):
                            counter += 1
                            if counter == 4:break
                            continue
                        else:mark = True
                    else:
                        if " р." in r.decode('ibm866'):
                            try:
                                postIndex = int((re.search('[0-9]{5}',r.decode('ibm866'))).group(0))
                                #print (postIndex)
                            except AttributeError:
                                postIndex = None
                            if postIndex and postIndex in self.indexList:
                                continue
                            elif not postIndex:
                                postName = r.decode('ibm866')[r.decode('ibm866').index(" р.") + 1:].rstrip().replace(' ','')
                                postNameTuple = tuple(postName[2:].split('-'))
                                #print("No index in post name,file----> ", postName,filename)
                                yield (postNameTuple, postName, filename, postIndex, year)
                        elif re.search('[Кк]анал',r.decode('ibm866')):
                            try:
                                postIndex = int((re.search('[0-9]{5}',r.decode('ibm866'))).group(0))
                            except AttributeError:
                                postIndex = None
                            if postIndex and postIndex in self.indexList:
                                continue
                            elif not postIndex:                                                         # None not in self.indexList and postIndex == None or None in self.indexList:
                                postName = r.decode('ibm866')[r.decode('ibm866').index("анал") + 1:].rstrip().replace(' ','')
                                postNameTuple = tuple(postName[3:].split('-'))
                                #print("No index in post name,file----> ", postName,filename)
                                yield (postNameTuple, postName, filename, postIndex, year)

def parseDocRowList(filenameGenerator):
	for_one_year_list = []
	filenameGenerator = list(filenameGenerator)
	count = 1
	for filenameyear in filenameGenerator:
	    if len(for_one_year_list) == 0:
	        for_one_year_list.append(filenameyear[1])
	    elif filenameyear[1] not in for_one_year_list and len(for_one_year_list) != 0:
	        yield for_one_year_list
	        for_one_year_list = []
	        for_one_year_list.append(filenameyear[1])
	    with open(filenameyear[0], 'rb') as f:
	    	print(filenameyear[0])
	    	for r in f:
	    		if " р." in r.decode('ibm866'):
	    		    postIndex = int((re.search('[0-9]{5}',r.decode('ibm866'))).group(0))
	    		    #print (postIndex)
	    		    postName = r.decode('ibm866')[r.decode('ibm866').index(" р.") + 1:].rstrip().replace(' ','')
	    		    for_one_year_list.append(postName)
	    		    #postNameTuple = tuple(postName[2:].split('-'))
	    		    #print(repr(postNameTuple))
	    		    #yield (postNameTuple, postName, filename, postIndex, year)
	    		elif re.search('[Кк]анал',r.decode('ibm866')):
	    		    postIndex = int((re.search('[0-9]{5}',r.decode('ibm866'))).group(0))
	    		    postName = r.decode('ibm866')[r.decode('ibm866').index("анал") + 1:].rstrip().replace(' ','')
	    		    for_one_year_list.append(postName)
	    		    #postNameTuple = tuple(postName[3:].split('-'))
	    		    #print(repr(postNameTuple))
	    		    #yield (postNameTuple, postName, filename, postIndex, year)
	    	#print(for_one_year_list)
	    	if count == len(filenameGenerator):
	    		yield for_one_year_list
	    	else:
	    		count += 1
	    #yield for_one_year_list

def generate_file(confFile,XLS,isIncludeNoneIndex=False):
    listSHeetes = input("Please,choose massive of sheetes you want to work with.\nFor example,type: [5:11] or [5:] or [:11]  or just press ENTER to work with all sheetes.")
    wb = openpyxl.load_workbook(XLS)
    allSheetes = eval("wb.get_sheet_names()" + listSHeetes)
    print(allSheetes)
    print("\nTotal number of sheetes you've chosen = {}\n".format(len(allSheetes)))
    with open(confFile,"w") as f:
        dataDict = {}
        for sheet in allSheetes:
            sht = wb.get_sheet_by_name(sheet)
            try:
                index_obj = re.search(r'[0-9]{5}',str(sht.cell(row=2,column=1).value))
                index = int(index_obj.group(0))
                chapterOfSheet = str(sht.cell(row=2,column=1).value)[index_obj.end():].rstrip().replace(' ','')
            except AttributeError:
                if isIncludeNoneIndex:
                    index = None
                    chapterOfSheet = str(sht.cell(row=2,column=1).value).strip().replace(' ','')
                else:
                    continue
            print(chapterOfSheet)
            oneSheet = {sheet:[index, chapterOfSheet]}
            dataDict.update(oneSheet)
        print("\nNumber of names to work with = {}".format(len(dataDict)))
        json.dump(dataDict, f, ensure_ascii=False, indent=4)


def choose_path():
    root = tk.Tk()
    root.withdraw()
    root = filedialog.askdirectory(title="Select A Folder to Data Base", mustexist=1)
    return root

def choose_file(opts=opts):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(**opts)
    return file_path

def write_nonindex_names_to_conf(jsonData, jsonFile, root, obj, indexList, ratioOfDiff=2.55):
    write_posts_names_to_config(seeDiff(jsonData,obj.parseDoc(walking(root), indexList), ratioOfDiff), jsonData, jsonFile)

if __name__ == "__main__":
    # parseDocFilename = "T78281.V3"
    # parseDocFilename1 = "Tabl1.12 sivDon"
    # parseDocFilename2 = "T78281.txt"
    #seeDiff("configure.json",parseDoc(parseDocFilename2))
    root = choose_path()
    if root:
        jsonData = load_json_conf("configure.json")
        indexList = [jsonData[x][0] for x in jsonData]
        #write_index_year_table(parseDocRowList(walking(root)))
        createObj = createConf()
        createObj.add_table_number_pattern('таблиц[ая] ?1.12')
        createObj.add_table_name_pattern('температура вод[иы]')
        write_nonindex_names_to_conf(createObj)
        print(indexList)
        print("Chosen path to dataBase is: ",root)
    print("Exit.")
    #seeDiff("configure.json")
    # for i in parseDoc("Tabl1.12 sivDon"):
    # 	print(i)