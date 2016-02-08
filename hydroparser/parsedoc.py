#!/usr/bin/env python
# -*- coding: utf8 -*-
import openpyxl
import re
import os
import sys
import json
from tkinter import filedialog
import tkinter as tk


def walkingDead(pathToFiles,docObject):
    for folder in os.listdir(pathToFiles):
        try:
            folderInteger = int(folder)
        except ValueError:
            folderInteger = None
        print ("--",folderInteger,folder)
        for dirPath, dirs, files in os.walk(os.path.join(pathToFiles,folder)):
            for fileSingle in files:
                if fileSingle[-4:] == 'xlsx':
                    print(os.path.join(dirPath,fileSingle))
                    for gen in docObject.parse_XLSX(os.path.join(dirPath,fileSingle),folderInteger):
                        yield gen
                else:
                    for gen in docObject.parse_doc(os.path.join(dirPath,fileSingle),folderInteger):
                        yield gen  

class ParseOneDoc:

    def __init__(self, jsonData=''):
        # self.filename = ''
        # self.year = None
        self.jsonData = jsonData
        self.table_number_pattern = 'таблиц[ая] ?1.12'
        self.table_name_pattern = 'температура вод[иы]'
        self.data_marker = '[СCсc]е?редн.'
        self.marker_of_list_pattern = 0

    def add_table_number_pattern(self, table_number_pattern):
        self.table_number_pattern = table_number_pattern
        #return self.table_number_pattern

    def add_table_name_pattern(self, table_name_pattern):
        self.table_name_pattern = table_name_pattern
        #return self.table_name_pattern

    def add_data_marker(self, data_marker):
        self.data_marker = data_marker
        #return self.data_marker

    def select_data_list_pattern(self, marker_of_list_pattern):
        self.marker_of_list_pattern = marker_of_list_pattern
        #return self.marker_of_list_pattern

    def parse_doc(self, filename, year):
        self.filename = filename
        self.year = year
        with open(self.filename, 'rb') as f:
            flag = False
            postName = ''
            mark = False
            counter = 0
            for r in f:
                if not mark:
                    if not re.search(self.table_number_pattern,r.decode('ibm866').lower()) and not re.search(self.table_name_pattern,r.decode('ibm866').lower()):
                        counter += 1
                        if counter == 4:break
                        continue
                    else:
                        if not self.year:
                            try:
                                self.year = int((re.search(r'[0-9]{4}',r.decode('ibm866'))).group(0))
                            except AttributeError:
                                pass
                            except TypeError:
                                pass
                        mark = True
                else:
                    if " р." in r.decode('ibm866') or re.search('[Кк]анал',r.decode('ibm866')):
                        try:
                            postName = r.decode('ibm866')[r.decode('ibm866').index(" р.")+1:].rstrip()
                        except ValueError:
                            postName = r.decode('ibm866')[r.decode('ibm866').index("анал"):].rstrip()
                        try:
                            postIndex = int((re.search('[0-9]{5}',r.decode('ibm866'))).group(0))
                        except AttributeError:
                            postIndex = None
                        flag = True
                    if re.search(self.data_marker,r.decode('ibm866')) and flag:
                        if self.marker_of_list_pattern == 0:
                            li = [x for x in r.decode('ibm866').lstrip().split()[1:13]]
                        
                        flag = False
                        if postIndex:
                            for mainKey in self.jsonData:
                                if postIndex == self.jsonData[mainKey][0]:
                                    yield (mainKey,li,self.year,postIndex)
                        else:
                            for mainKey in self.jsonData:
                                if postName.replace(' ','') in self.jsonData[mainKey] or postName in self.jsonData[mainKey]:
                                    #postName = postName.replace(' ','')
                                    yield (mainKey,li,self.year,postIndex)

    def parse_XLSX(self, filename,year):
        self.filename = filename
        self.year = year
        postIndex = None
        try:
            wb = openpyxl.load_workbook(self.filename)
            try:
                sheet = wb.worksheets[5]
                print ("PARSE ",self.filename,self.year)
                catchIt = False
                try:
                    for i in range(1,4):
                        try:
                            if re.search(self.table_number_pattern,sheet.cell(row=i,column=1).value.lower()):
                                catchIt = True
                                if not year:
                                    for index in range(1,sheet.get_highest_column() + 1):
                                        try:
                                            postIndex = int((re.search(r'[0-9]{5}',sheet.cell(row=i,column=index).value)).group(0))
                                            if postIndex:break
                                        except (AttributeError, TypeError):
                                            try:
                                                year = int((re.search(r'[0-9]{4}',sheet.cell(row=i,column=index).value)).group(0))
                                                print (type(year))
                                                if year:break
                                            except (AttributeError, TypeError):
                                                pass
                                print (filename)
                                break
                            else:sheet = None
                        except AttributeError:
                            pass
                except TypeError:
                    sheet = None
            except IndexError:
                for name in wb.get_sheet_names():
                    sheet = wb.get_sheet_by_name(name)
                    catchIt = False
                    try:
                        for i in range(1,4):
                            try:
                                if re.search(self.table_number_pattern,sheet.cell(row=i,column=1).value.lower()):
                                    catchIt = True
                                    if not year:
                                        for index in range(1,sheet.get_highest_column() + 1):
                                            try:
                                                postIndex = int((re.search(r'[0-9]{5}',sheet.cell(row=i,column=index).value)).group(0))
                                                if postIndex:break
                                            except (AttributeError, TypeError):
                                                try:
                                                    year = int((re.search(r'[0-9]{4}',sheet.cell(row=i,column=index).value)).group(0))
                                                    if year:break
                                                except (AttributeError, TypeError):
                                                    pass
                                    break
                                else:sheet = None
                            except AttributeError:
                                pass
                        if catchIt:break
                    except TypeError:
                        sheet = None
        except TypeError:
            print ("Can't load_workbook..")
            sheet = None
        except:
            print("Unexpected error:", sys.exc_info()[0])
            sheet = None

        if sheet:
            if postIndex:
                for i in range(2,sheet.get_highest_row()+1):
                    try:
                        if re.search(r'[0-9]{4}',str(sheet.cell(row=i,column=1).value)) and str(postIndex) not in str(sheet.cell(row=i,column=1).value):
                            year = int(re.search(r'[0-9]{4}',str(sheet.cell(row=i,column=1).value)).group(0))
                            if self.marker_of_list_pattern == 0:
                                li = [sheet.cell(row=i+4,column=x).value for x in range(4,16) ]
                            for mainKey in self.jsonData:
                                if postIndex == self.jsonData[mainKey][0]:
                                    yield (mainKey,li,year,postIndex)
                        if re.search(r'[0-9]{5}',str(sheet.cell(row=i,column=1).value)) and str(postIndex) not in str(sheet.cell(row=i,column=1).value):
                            postIndex = int(re.search(r'[0-9]{5}',str(sheet.cell(row=i,column=1).value)).group(0))
                            print("NEW postIndex is:",postIndex)
                    except TypeError:
                        pass
            else:
                for i in range(1,sheet.get_highest_row()+1):
                    try:
                        if " р." in sheet.cell(row=i,column=1).value or re.search('[Кк]анал', sheet.cell(row=i,column=1).value):
                            try:
                                postName = sheet.cell(row=i,column=1).value[sheet.cell(row=i,column=1).value.index(" р.")+1:].rstrip()
                            except ValueError:
                                postName = sheet.cell(row=i,column=1).value[sheet.cell(row=i,column=1).value.index("анал"):].rstrip()
                            try:
                                postIndex = int((re.search('[0-9]{5}',sheet.cell(row=i,column=1).value)).group(0))
                            except AttributeError:
                                postIndex = None
                            if self.marker_of_list_pattern == 0:
                                li = [sheet.cell(row=i+4,column=x).value for x in range(4,16) ]
                            if postIndex:
                                for mainKey in self.jsonData:
                                    if postIndex == self.jsonData[mainKey][0]:
                                        yield (mainKey,li,year,postIndex)
                            else:
                                for mainKey in self.jsonData:
                                    if postName.replace(' ','') in self.jsonData[mainKey] or postName in self.jsonData[mainKey]:
                                        yield (mainKey,li,year,postIndex)

                    except TypeError:
                        pass


def updateXLS(sheetYearData, XLSfile):                       # sheetYearData  must be a generator
    wb = openpyxl.load_workbook(XLSfile)
    for row in sheetYearData:
        sheet = wb.get_sheet_by_name(row[0])
        for index in range(5,sheet.get_highest_row()+1):
            yearOfSheet = sheet.cell(row=index, column=1).value
            if yearOfSheet == row[2]:
                for i in range(len(row[1])):
                    try:
                        sheet.cell(row=index, column=i+2).value = float(row[1][i])
                    except ValueError:
                        sheet.cell(row=index, column=i+2).value = row[1][i].replace('-','')
    wb.save(XLSfile)


# def Choose_File():
#     opts = {}
#     opts['filetypes'] = [('XLSX files','.xlsx'), ('All files', '.*')]
#     opts['title'] = 'Select XLSX file for inserting datas'
#     root = tk.Tk()
#     root.withdraw()
#     file_path = filedialog.askopenfilename(**opts)
#     return file_path

if __name__ == "__main__":
    with open("configure.json") as f:
        jsonData = json.load(f)
    XLSfile = Choose_File()
    pathToFiles = filedialog.askdirectory(title="Select A Folder", mustexist=1)
    if XLSfile and pathToFiles:
        print(repr(XLSfile),pathToFiles,len(XLSfile),len(pathToFiles))
        temperatureObject = ParseOneDoc()
        temperatureObject.add_data_marker('[СCсc]е?редн.')
        temperatureObject.add_table_number_pattern('таблиц[ая] ?1.12')
        temperatureObject.add_table_name_pattern('температура вод[иы]')
        updateXLS(walkingDead(pathToFiles, temperatureObject),XLSfile)
        #updateIndexXLS(jsonData,XLSfile)
        #walkingDead(pathToFiles)
        #updateXLS(parseDoc("Tabl1.12 sivDon"),2008,"Seredniomisyachna_temper_vodi.xlsx","СівДон-Печ")
        # parseDoc("T78281.V3")
        # parseDoc("Tabl1.12 sivDon")
    #parseXLSX('/home/sashko/Python3/Programming/parseDataFromDoc/Files/2012/2012_V3_T_1_8_9_10_11_12_13_14.xlsx')
    print("Goodbye!Exit.")

