#!/usr/bin/env python
# -*- coding: utf8 -*-

import re
import openpyxl
import xlsxwriter
import math
import statistics

class Charter:
    def __init__(self,xlsxFile):
        self.xlsxFile = xlsxFile
        self.setSheetes = []
        self.yearList = []

    def set_range_of_sheets(self):
        rangeSheetes = input("Please,choose massive of sheetes you want to work with.\nFor example,type: [5:11] or [5:] or [:11]  or just press ENTER to work with all sheetes.")
        wb = openpyxl.load_workbook(self.xlsxFile)
        self.setSheetes = eval("wb.get_sheet_names()" + rangeSheetes)
        if not isinstance(self.setSheetes,list):
            self.setSheetes = [self.setSheetes]
        print(self.setSheetes)
        print("\nTotal number of sheetes you've chosen = {}\n".format(len(self.setSheetes)))

    def parseXLSXfile(self):
        wb = openpyxl.load_workbook(self.xlsxFile)
        sh = wb.worksheets[0]
        if len(self.setSheetes) == 0:self.setSheetes = wb.get_sheet_names()
        xlsxGeneratorList = []
        for sheetName in self.setSheetes:
            sheet = wb.get_sheet_by_name(sheetName)
            sheetDict = {}
            highestrow = sheet.max_row
            gutIt = False
            findFirstAvarege = False
            for i in range(4,highestrow+1):
                try:
                    if re.search(r'[0-9]{4}',str(sheet.cell(row=i,column=1).value)):
                        yearInt = int(sheet.cell(row=i,column=1).value)
                        number = 0
                        gutIt = True
                        try:
                            avarege = statistics.mean([float(sheet.cell(row=i,column=x).value) for x in range(2,14)])
                            findFirstAvarege = True
                        except (TypeError,ValueError):
                            avarege = None 
                        if findFirstAvarege: 
                            sheetDict.update({yearInt : avarege})
                            if yearInt not in self.yearList: self.yearList.append(yearInt)
                except (TypeError, ValueError):
                    pass
            if gutIt:
                xlsxGeneratorList.append((sheetName,sheetDict))
        self.yearList.sort()
        return xlsxGeneratorList

    def correcting_data(self,xlsxGenerator):
        for sheet in xlsxGenerator:
            header = sheet[0]
            sheetDict = sheet[1]
            for i in self.yearList:
                if i not in  sheetDict:
                    sheetDict.update({i:None})
            yield (header,sheetDict)

    def prepareToMethod(self,xlsxGenerator):
        for sheet in xlsxGenerator:
            header = sheet[0]
            sheetDict = sheet[1]
            listOfdata = []
            for k in self.yearList:
                listOfdata.append(sheetDict[k])
            yield (header,listOfdata)
        
    

    def prepareToChart(self,Years,listOfDataheader):
        data = []
        headings = []
        for header,values in listOfDataheader:
            if len(headings) == 0 and len(data) == 0:
                headings.append('Years')
                data.append(Years)
            headings.append(header)
            data.append(values)
        return(headings,data)




def createCharts(methodsList,xlsxFile):
    inputXLSXfile = xlsxFile[:-5] + "_Charts.xlsx"
    workbook = xlsxwriter.Workbook(inputXLSXfile)
    for chartData,sheetName,private in methodsList:
        worksheet = workbook.add_worksheet(sheetName)
        print ("Sheet {} has been created.".format(sheetName))
        bold = workbook.add_format({'bold': 1})
        # Add the worksheet data that the charts will refer to.
        headings = chartData[0]
        data = chartData[1]
        worksheet.write_row('A1', headings, bold)
        for i,valueList in enumerate(data):
            for r,value in enumerate(valueList):
                if isinstance(value,(float,int)) and not math.isnan(value):
                    worksheet.write(r + 1, i, value)
        # Create a new chart object. In this case an embedded chart.
        chart1 = workbook.add_chart({'type': 'line'})
        chart2 = workbook.add_chart({'type': 'line'})
        # Configure the series.
        rowdistance = len(data[0])+5
        coldistance = 0
        count = 1
        for i in range(1, len(data)):
            if private == 1 or private == 2:
                chart1 = workbook.add_chart({'type': 'line'})
                chart1.add_series({
                'name':       [sheetName, 0, i],
                'categories': [sheetName, 1, 0, len(data[0]), 0],
                'values':     [sheetName, 1, i, len(data[0]), i],
                'marker': {'type': 'diamond'},
                'trendline': {'type': 'linear'},
                })
                chart1.set_size({'width': 700, 'height': 500})
                # Add a chart title and some axis labels.
                chart1.set_title ({'name': headings[i]})
                chart1.set_x_axis({'name': headings[0]})
                chart1.set_y_axis({'name': 'Something'})
                # Set an Excel chart style. Colors with white outline and shadow.
                chart1.set_style(10)
                # Insert the chart into the worksheet (with an offset).
                worksheet.insert_chart(rowdistance, coldistance, chart1, {'x_offset': 25, 'y_offset': 10})
                if count % 7 == 0:
                    rowdistance += 30
                    coldistance = 0
                else:
                    coldistance += 12
                count += 1
            if private == 0 or private == 2:
                chart2.add_series({
                'name':       [sheetName, 0, i],
                'categories': [sheetName, 1, 0, len(data[0]), 0],
                'values':     [sheetName, 1, i, len(data[0]), i],
                'marker': {'type': 'diamond'},
            })
        if private == 0 or private  == 2:
            chart2.set_size({'width': 2000, 'height': 1000})
            # Add a chart title and some axis labels.
            chart2.set_title ({'name': sheetName})
            chart2.set_x_axis({'name': headings[0]})
            chart2.set_y_axis({'name': 'Samething'})
            # Set an Excel chart style. Colors with white outline and shadow.
            chart2.set_style(10)
            # Insert the chart into the worksheet (with an offset).
            worksheet.insert_chart(0,len(headings), chart2, {'x_offset': 25, 'y_offset': 10})

    workbook.close()
    print ("File {} has been saved.".format(inputXLSXfile))


class GydroGenMethod:
    """Using Gydro Genethich method in flow calculation"""
    empcount = 0

    def __init__(self,Q):
        self.Q = Q
        GydroGenMethod.empcount += 1

    def rizn_inter_kruvi(self):
        calculate_list = []
        Qfllist = [x for x in self.Q if isinstance(x, float)]
        lenQfllist = len(Qfllist)
        Qav = statistics.mean(Qfllist)
        sumQ = math.fsum([(Qf - Qav)**2 for Qf in Qfllist])
        lastValue = None
        for i,q in enumerate(self.Q):
            if not lastValue and q:
                lastValue = q/Qav - 1
                calculate_list.append(lastValue)
                continue
            elif not q:
                calculate_list.append(None)
            else:
                calculate_list.append(q/Qav - 1 + lastValue)
                lastValue += q/Qav - 1
        y = math.sqrt(sumQ/(lenQfllist-1))/Qav
        rizn_int_kr = [float("{0:.2f}".format(x/y)) if x is not None else None for x in calculate_list]
        return rizn_int_kr

    def sumCurves(self):
        curveList = []
        lastValue = None
        for i,value in enumerate(self.Q):
            if not lastValue and value:
                curveList.append(value)
                lastValue = value
            elif lastValue and value:
                curveList.append(value + lastValue)
                lastValue += value
            else:
                curveList.append(None)
        return curveList

def build_charts(xlsxFile):
    outputFile = Charter(xlsxFile)
    outputFile.set_range_of_sheets()
    genFile = outputFile.parseXLSXfile()
    newGenFile = outputFile.correcting_data(genFile)
    prepareForMethods = outputFile.prepareToMethod(newGenFile)
    listOfDataHeader = []
    listOfSumCurvesHeader = []
    listOfChronology = []
    for f in prepareForMethods:
        riznIntKr = GydroGenMethod(f[1])
        rizn = riznIntKr.rizn_inter_kruvi()
        listOfDataHeader.append((f[0],rizn))
        sumCurve = riznIntKr.sumCurves()
        listOfSumCurvesHeader.append((f[0],sumCurve))
        listOfChronology.append((f[0],f[1]))
    years = outputFile.yearList
    indexOfYears = range(len(years))
    MethodRizInt = outputFile.prepareToChart(years,listOfDataHeader)
    MethodSumCurves = outputFile.prepareToChart(indexOfYears,listOfSumCurvesHeader)
    MethodChronology = outputFile.prepareToChart(years,listOfChronology)
    methodsList = [(MethodRizInt,"Rizn_Int_Kruvi",2),(MethodSumCurves,"Sum Curves",1),(MethodChronology,"Chronology",1)]
    createCharts(methodsList,xlsxFile)



if __name__ == "__main__":
    build_charts(xlsxFile)
    #print (prepare(parseXLSXfile(xlsxFile),"rizn_inter_kruvi"))
    #createCharts(prepare(parseXLSXfile(xlsxFile),"rizn_inter_kruvi"))