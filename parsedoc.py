#!/usr/bin/env python
# -*- coding: utf8 -*-
import openpyxl
import re
import os
import sys
import json
from tkinter import filedialog
import tkinter as tk


def walkingDead(pathToFiles='.'):
    for folder in os.listdir(pathToFiles):
        try:
            folderInteger = int(folder)
        except ValueError:
            folderInteger = None
        print ("--",folderInteger)
        for dirPath, dirs, files in os.walk(os.path.join(pathToFiles,folder)):
            for fileSingle in files:
                if fileSingle[-4:] == 'xlsx':
                    for gen in parseXLSX(os.path.join(dirPath,fileSingle),folderInteger):
                        yield gen
                else:
                    for gen in parseDoc(os.path.join(dirPath,fileSingle),folderInteger):
                        yield gen  

def parseDoc(filename, year):
    with open(filename, 'rb') as f:
        flag = False
        postName = ''
        mark = False
        counter = 0
        for r in f:
            if not mark:
                if not re.search('[Тт]аблиц[ая] ?1.12',r.decode('ibm866')) and not re.search('[Тт]емпература вод[иы]',r.decode('ibm866')):
                    counter += 1
                    if counter == 4:break
                    continue
                else:
                    if not year:
                        try:
                            year = int((re.search(r'[0-9]{4}',r.decode('ibm866'))).group(0))
                        except AttributeError:
                            pass
                        except TypeError:
                            pass
                    mark = True
            else:
                if " р." in r.decode('ibm866') or re.search('[Кк]анал',r.decode('ibm866')):
                    try:
                        postName = r.decode('ibm866')[r.decode('ibm866').index(" р.")+1:].rstrip()
                    except ValueError:
                        postName = r.decode('ibm866')[r.decode('ibm866').index("анал"):].rstrip()
                    try:
                        postIndex = int((re.search('[0-9]{5}',r.decode('ibm866'))).group(0))
                    except AttributeError:
                        postIndex = None
                    flag = True
                if re.search('[Сс]е?редн.',r.decode('ibm866')) and flag:
                    li = [x for x in r.decode('ibm866').lstrip().split()[1:13]]
                    #li = [float(x) if re.match("^\d+?\.\d+?$", x) is not None else "" for x in r.decode('ibm866').lstrip().split()[1:13]]
                    flag = False
                    if postIndex:
                        for mainKey in jsonData:
                            if postIndex == jsonData[mainKey][0]:
                                yield (mainKey,li,year,postIndex)
                    else:
                        for mainKey in jsonData:
                            if postName.replace(' ','') in jsonData[mainKey] or postName in jsonData[mainKey]:
                                #postName = postName.replace(' ','')
                                yield (mainKey,li,year,postIndex)

def parseXLSX(filename,year):
    print(filename)
    try:
        wb = openpyxl.load_workbook(filename)
        try:
            sheet = wb.worksheets[5]
            print ("PARSE ",filename,year)
            try:
                if re.search('ТАБЛИЦ[АЯ] ?1.12',sheet.cell(row=1,column=1).value):
                    if not year:
                        for index in range(1,sheet.get_highest_column() + 1):
                            try:
                                year = int((re.search(r'[0-9]{4}',sheet.cell(row=1,column=index).value)).group(0))
                                print (type(year))
                                if year:break
                            except (AttributeError, TypeError):
                                pass
                            # except TypeError:
                            #     pass
                    print (filename)
                else:sheet = None
            except TypeError:
                sheet = None
        except IndexError:
            for name in wb.get_sheet_names():
                sheet = wb.get_sheet_by_name(name)
                try:
                    if re.search('ТАБЛИЦ[АЯ] ?1.12',sheet.cell(row=1,column=1).value):
                        if not year:
                            for index in range(1,sheet.get_highest_column() + 1):
                                try:
                                    year = int((re.search(r'[0-9]{4}',sheet.cell(row=1,column=index).value)).group(0))
                                    if year:break
                                except AttributeError:
                                    pass
                                except TypeError:
                                    pass
                        break
                    else:sheet = None
                except TypeError:
                    sheet = None
    except TypeError:
        print ("Can't load_workbook..")
        sheet = None
    except:
        print("Unexpected error:", sys.exc_info()[0])
        sheet = None

    if sheet:
        for i in range(1,sheet.get_highest_row()+1):
            #print (sheet.cell(row=i,column=1).value)
            try:
                if " р." in sheet.cell(row=i,column=1).value or re.search('[Кк]анал', sheet.cell(row=i,column=1).value):
                    try:
                        postName = sheet.cell(row=i,column=1).value[sheet.cell(row=i,column=1).value.index(" р.")+1:].rstrip()
                    except ValueError:
                        postName = sheet.cell(row=i,column=1).value[sheet.cell(row=i,column=1).value.index("анал"):].rstrip()
                    try:
                        postIndex = int((re.search('[0-9]{5}',sheet.cell(row=i,column=1).value)).group(0))
                    except AttributeError:
                        postIndex = None
                    li = [sheet.cell(row=i+4,column=x).value for x in range(4,16) ]
                    #li = [float(sheet.cell(row=i+4,column=x).value) if re.search("\d+?\.\d+?", sheet.cell(row=i+4,column=x).value) is not None else "" for x in range(4,16) ]
                    #print (postName,postIndex)
                    #print (li)
                    if postIndex:
                        for mainKey in jsonData:
                            if postIndex == jsonData[mainKey][0]:
                                yield (mainKey,li,year,postIndex)
                    else:
                        for mainKey in jsonData:
                            if postName.replace(' ','') in jsonData[mainKey] or postName in jsonData[mainKey]:
                                #postName = postName.replace(' ','')
                                yield (mainKey,li,year,postIndex)
                    # print (float(sheet.cell(row=i+4,column=4).value))
                    # print (re.search("\d+?\.\d+?", sheet.cell(row=i+4,column=4).value))
            except TypeError:
                pass


def updateXLS(sheetYearData, XLSfile):                       # sheetYearData  must be a generator
    wb = openpyxl.load_workbook(XLSfile)
    for row in sheetYearData:
        sheet = wb.get_sheet_by_name(row[0])
        for index in range(5,sheet.get_highest_row()+1):
            yearOfSheet = sheet.cell(row=index, column=1).value
            if yearOfSheet == row[2]:
                for i in range(len(row[1])):
                    try:
                        sheet.cell(row=index, column=i+2).value = float(row[1][i])
                    except ValueError:
                        sheet.cell(row=index, column=i+2).value = row[1][i]
    wb.save(XLSfile)

def updateIndexXLS(jsonData, XLSfile):
    wb = openpyxl.load_workbook(XLSfile)
    for mainKey in jsonData:
        print(mainKey)
        if not jsonData[mainKey][list(jsonData[mainKey].keys())[0]][0]:
            print (jsonData[mainKey])
            try:
                index = jsonData[mainKey][list(jsonData[mainKey].keys())[0]][1][1]
                print (index)
                sheet = wb.get_sheet_by_name(mainKey)
                sheet.cell(row=2,column=1).value = str(index) + ' ' + sheet.cell(row=2,column=1).value
            except IndexError:
                pass
    wb.save(XLSfile)


def choose_file():
    opts = {}
    opts['filetypes'] = [('XLSX files','.xlsx'), ('All files', '.*')]
    opts['title'] = 'Select XLSX file for inserting datas'
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(**opts)
    return file_path

if __name__ == "__main__":
    with open("configure.json") as f:
        jsonData = json.load(f)
    XLSfile = choose_file()
    pathToFiles = filedialog.askdirectory(title="Select A Folder", mustexist=1)
    if XLSfile and pathToFiles:
        print(repr(XLSfile),pathToFiles,len(XLSfile),len(pathToFiles))
        updateXLS(walkingDead(pathToFiles),XLSfile)
        #updateIndexXLS(jsonData,XLSfile)
        #walkingDead(pathToFiles)
        #updateXLS(parseDoc("Tabl1.12 sivDon"),2008,"Seredniomisyachna_temper_vodi.xlsx","СівДон-Печ")
        # parseDoc("T78281.V3")
        # parseDoc("Tabl1.12 sivDon")
    #parseXLSX('/home/sashko/Python3/Programming/parseDataFromDoc/Files/2012/2012_V3_T_1_8_9_10_11_12_13_14.xlsx')
    print("Goodbye!Exit.")

